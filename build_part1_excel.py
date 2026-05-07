"""
Build the deliverable 'Results for Part I-SAAM.xlsx' from the cleaned Part I outputs.
This populates the template (Annualized average return, vol, cumulative return,
Sharpe ratio, min, max) for the value-weighted and minimum-variance portfolios,
inserts the 144 monthly returns from 2014-01 to 2025-12, and embeds the cumulative
performance plot.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage


PROJECT_DIR = Path(__file__).resolve().parent
TEMPLATE = PROJECT_DIR / "outputs_part1" / "Template for Part I-SAAM (1).xlsx"
OUTPUT = PROJECT_DIR / "outputs_part1" / "Results for Part I-SAAM.xlsx"
STATS_CSV = PROJECT_DIR / "outputs_part1" / "part1_summary_stats.csv"
RETURNS_CSV = PROJECT_DIR / "outputs_part1" / "part1_monthly_returns_2014_2025.csv"
PLOT_PNG = PROJECT_DIR / "outputs_part1" / "part1_cumulative_returns.png"


def main() -> None:
    stats = pd.read_csv(STATS_CSV).set_index("portfolio")
    returns = pd.read_csv(RETURNS_CSV, parse_dates=["date"]).set_index("date")
    returns = returns.sort_index()

    wb = load_workbook(TEMPLATE)
    ws = wb["Sheet1"]

    vw_col = "B"
    mv_col = "C"

    def fill(row: int, vw_value: float, mv_value: float) -> None:
        ws[f"{vw_col}{row}"] = float(vw_value)
        ws[f"{mv_col}{row}"] = float(mv_value)

    # Row 3: Annualized average return (arithmetic)
    fill(3, stats.loc["P_oos_vw", "ann_return_arith"], stats.loc["P_oos_mv", "ann_return_arith"])
    # Row 4: Annualized volatility
    fill(4, stats.loc["P_oos_vw", "ann_vol"], stats.loc["P_oos_mv", "ann_vol"])
    # Row 5: Annualized cumulative return (geometric)
    fill(5, stats.loc["P_oos_vw", "ann_return_geom"], stats.loc["P_oos_mv", "ann_return_geom"])
    # Row 6: Sharpe ratio (arithmetic Sharpe = arith ann return / ann vol)
    fill(6, stats.loc["P_oos_vw", "sharpe_arith"], stats.loc["P_oos_mv", "sharpe_arith"])
    # Row 7: Minimum monthly return
    fill(7, stats.loc["P_oos_vw", "min"], stats.loc["P_oos_mv", "min"])
    # Row 8: Maximum monthly return
    fill(8, stats.loc["P_oos_vw", "max"], stats.loc["P_oos_mv", "max"])

    # Monthly returns: Column E = date (already filled by template), F = VW, G = MV.
    # Iterate template date column starting at row 3 to row n.
    template_dates = []
    for row_idx in range(3, ws.max_row + 1):
        cell_value = ws[f"E{row_idx}"].value
        if cell_value is None:
            break
        template_dates.append((row_idx, pd.Timestamp(cell_value)))

    returns_periodic = returns.copy()
    returns_periodic["period"] = returns_periodic.index.to_period("M")
    returns_by_period = returns_periodic.set_index("period")

    for row_idx, date_val in template_dates:
        period_key = pd.Period(date_val, freq="M")
        if period_key in returns_by_period.index:
            row = returns_by_period.loc[period_key]
            ws[f"F{row_idx}"] = float(row["R_vw"]) if pd.notna(row["R_vw"]) else None
            ws[f"G{row_idx}"] = float(row["R_mv"]) if pd.notna(row["R_mv"]) else None

    # Insert the cumulative returns plot below the metrics block.
    if PLOT_PNG.exists():
        img = XLImage(str(PLOT_PNG))
        img.width = 720
        img.height = 360
        ws.add_image(img, "A12")

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Saved Excel deliverable to: {OUTPUT}")


if __name__ == "__main__":
    main()
